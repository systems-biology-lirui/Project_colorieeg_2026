"""Build a transparent signal/localization intersection manifest."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from .config import (
    KNOWN_BAD_CHANNELS,
    MANUAL_DECISIONS_PATH,
    PROTECTED_CHANNELS,
    QC_ROOT,
    SUBJECTS,
    TASKS,
    SOURCE_METADATA_ROOT,
)
from .io_seeg import load_set_metadata


def _natural_key(label: str) -> tuple[str, int, str]:
    text = label.strip().upper()
    match = re.fullmatch(r"([A-Z]+)(\d+)", text)
    if match:
        return match.group(1), int(match.group(2)), text
    numeric = re.fullmatch(r"\d+", text)
    if numeric:
        return "ZZZ", int(text), text
    return text, -1, text


def _clean_label(value: object) -> str:
    return str(value).strip().upper().replace(" ", "")


def localization_path(subject: str) -> Path:
    directory = SOURCE_METADATA_ROOT / subject
    xlsx = directory / f"{subject}_ieegloc.xlsx"
    if xlsx.exists():
        return xlsx
    tsv = directory / f"{subject}.tsv"
    if tsv.exists():
        return tsv
    raise FileNotFoundError(f"No localization table found for {subject}: {xlsx} or {tsv}")


def load_localized_labels(subject: str) -> set[str]:
    """Read only channel labels from the subject localization table."""

    path = localization_path(subject)
    if path.suffix.lower() == ".tsv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle, delimiter="\t")
            if not reader.fieldnames or "Channel" not in reader.fieldnames:
                raise ValueError(f"Localization TSV has no Channel column: {path}")
            return {
                _clean_label(row.get("Channel", ""))
                for row in reader
                if _clean_label(row.get("Channel", ""))
            }

    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required to read localization xlsx files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return set()
    headers = [str(value).strip() if value is not None else "" for value in header]
    try:
        channel_index = headers.index("Channel")
    except ValueError as exc:
        raise ValueError(f"Localization xlsx has no Channel column: {path}") from exc
    labels: set[str] = set()
    for row in rows:
        if channel_index >= len(row):
            continue
        label = _clean_label(row[channel_index])
        if label:
            labels.add(label)
    return labels


def _as_bool(value: object) -> bool:
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是"}


def load_localization_annotations(subject: str) -> dict[str, dict[str, object]]:
    """Load optional stimulation annotations keyed by contact label.

    Duplicate rows are collapsed conservatively: boolean evidence is merged
    with OR and text evidence is joined. Atlas localization columns are not
    interpreted here.
    """

    path = localization_path(subject)
    rows: list[dict[str, object]] = []
    if path.suffix.lower() == ".tsv":
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            rows = [dict(row) for row in csv.DictReader(handle, delimiter="\t")]
    else:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ImportError("openpyxl is required to read localization xlsx files") from exc
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        header = next(iterator, None)
        if header is not None:
            headers = [str(value).strip() if value is not None else "" for value in header]
            rows = [dict(zip(headers, row)) for row in iterator]

    annotations: dict[str, dict[str, object]] = {}
    for row in rows:
        channel = _clean_label(row.get("Channel", ""))
        if not channel:
            continue
        current = annotations.setdefault(
            channel,
            {
                "color_with_sti": False,
                "stim_behavior_recorded": False,
                "stim_color_evidence": set(),
                "stim_behavior_pairs": set(),
            },
        )
        current["color_with_sti"] = bool(current["color_with_sti"]) or _as_bool(
            row.get("color_with_sti", False)
        )
        current["stim_behavior_recorded"] = bool(current["stim_behavior_recorded"]) or _as_bool(
            row.get("stim_behavior_recorded", False)
        )
        for field in ("stim_color_evidence", "stim_behavior_pairs"):
            value = str(row.get(field) or "").strip()
            if value:
                current[field].update(item for item in value.split(";") if item)
    for annotation in annotations.values():
        annotation["stim_color_evidence"] = ";".join(sorted(annotation["stim_color_evidence"]))
        annotation["stim_behavior_pairs"] = ";".join(sorted(annotation["stim_behavior_pairs"]))
    return annotations


def load_confirmed_exclusions(review_csv: Path | None = None) -> dict[str, set[str]]:
    """Load explicit exclusions from the review table and durable decisions."""

    review_csv = review_csv or (QC_ROOT / "bad_channel_candidates.csv")
    excluded: dict[str, set[str]] = {subject: set() for subject in SUBJECTS}
    paths = (review_csv, MANUAL_DECISIONS_PATH)
    for path in paths:
        if not path.exists():
            continue
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            for row in csv.DictReader(handle):
                subject = str(row.get("subject", "")).strip()
                channel = _clean_label(row.get("channel", ""))
                decision = str(row.get("manual_decision", row.get("decision", ""))).strip().lower()
                if subject in excluded and channel and decision in {
                    "exclude", "bad", "confirmed_bad", "remove", "yes", "1", "true"
                }:
                    excluded[subject].add(channel)
    return excluded


def build_manifest_rows(
    subjects: tuple[str, ...] = SUBJECTS,
    tasks: tuple[int, ...] = TASKS,
    review_csv: Path | None = None,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    """Return subject-level and task-level manifest rows."""

    exclusions = load_confirmed_exclusions(review_csv)
    subject_rows: list[dict[str, object]] = []
    task_rows: list[dict[str, object]] = []
    for subject in subjects:
        localized = load_localized_labels(subject)
        localization_annotations = load_localization_annotations(subject)
        metadata_by_task = {task: load_set_metadata(subject, task) for task in tasks}
        signal_labels_by_task = {
            task: {label.upper() for label in metadata.labels}
            for task, metadata in metadata_by_task.items()
        }
        all_labels = set(localized)
        for labels in signal_labels_by_task.values():
            all_labels.update(labels)

        known_bad = {label.upper() for label in KNOWN_BAD_CHANNELS.get(subject, ())}
        confirmed_bad = exclusions.get(subject, set())
        protected = {label.upper() for label in PROTECTED_CHANNELS.get(subject, ())}
        localization_file = str(localization_path(subject))

        for channel in sorted(all_labels, key=_natural_key):
            tasks_present = [task for task in tasks if channel in signal_labels_by_task[task]]
            signal_any = bool(tasks_present)
            is_localized = channel in localized
            is_bad = channel in confirmed_bad
            reasons: list[str] = []
            if not signal_any:
                reasons.append("localization_only_no_signal")
            if not is_localized and signal_any:
                reasons.append("signal_without_localization")
            if is_bad:
                reasons.append("user_confirmed_bad_channel")
            if channel in known_bad and not is_bad:
                reasons.append("previous_bad_record_not_confirmed_here")
            if channel in protected:
                reasons.append("protected_channel")
            annotation = localization_annotations.get(channel, {})
            subject_rows.append(
                {
                    "subject": subject,
                    "channel": channel,
                    "localized": is_localized,
                    "signal_any": signal_any,
                    "tasks_present": ",".join(str(task) for task in tasks_present),
                    "confirmed_bad": is_bad,
                    "protected": channel in protected,
                    "color_with_sti": bool(annotation.get("color_with_sti", False)),
                    "stim_behavior_recorded": bool(annotation.get("stim_behavior_recorded", False)),
                    "stim_color_evidence": str(annotation.get("stim_color_evidence", "")),
                    "stim_behavior_pairs": str(annotation.get("stim_behavior_pairs", "")),
                    "analysis_center_eligible": bool(is_localized and signal_any and not is_bad),
                    "manifest_reason": ";".join(reasons),
                    "localization_file": localization_file,
                }
            )

            for task in tasks:
                signal_present = channel in signal_labels_by_task[task]
                task_rows.append(
                    {
                        "subject": subject,
                        "task_num": task,
                        "channel": channel,
                        "signal_present": signal_present,
                        "localized": is_localized,
                        "confirmed_bad": is_bad,
                        "protected": channel in protected,
                        "color_with_sti": bool(annotation.get("color_with_sti", False)),
                        "stim_behavior_recorded": bool(annotation.get("stim_behavior_recorded", False)),
                        "stim_color_evidence": str(annotation.get("stim_color_evidence", "")),
                        "stim_behavior_pairs": str(annotation.get("stim_behavior_pairs", "")),
                        "analysis_center_eligible": bool(is_localized and signal_present and not is_bad),
                        "manifest_reason": ";".join(
                            reason
                            for reason in (
                                []
                                if signal_present
                                else ["signal_missing_for_task"]
                            )
                            + ([] if is_localized else ["signal_without_localization"])
                            + (["user_confirmed_bad_channel"] if is_bad else [])
                        ),
                        "localization_file": localization_file,
                    }
                )
    return subject_rows, task_rows


def _write_rows(rows: Iterable[dict[str, object]], output_path: Path) -> None:
    rows = list(rows)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        return
    with output_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_manifests(
    output_root: Path,
    subjects: tuple[str, ...] = SUBJECTS,
    tasks: tuple[int, ...] = TASKS,
    review_csv: Path | None = None,
) -> tuple[Path, Path]:
    subject_rows, task_rows = build_manifest_rows(subjects, tasks, review_csv)
    subject_path = output_root / "electrode_manifest_by_subject.csv"
    task_path = output_root / "electrode_manifest_by_task.csv"
    _write_rows(subject_rows, subject_path)
    _write_rows(task_rows, task_path)
    return subject_path, task_path


def load_task_analysis_centers(manifest_csv: Path, subject: str, task_num: int) -> set[str]:
    """Read the localized, non-excluded center labels for one recording."""

    centers: set[str] = set()
    with manifest_csv.open("r", newline="", encoding="utf-8-sig") as handle:
        for row in csv.DictReader(handle):
            if (
                str(row.get("subject", "")).strip() == subject
                and int(row.get("task_num", -1)) == int(task_num)
                and str(row.get("analysis_center_eligible", "")).strip().lower()
                in {"true", "1", "yes"}
            ):
                centers.add(_clean_label(row.get("channel", "")))
    return centers
