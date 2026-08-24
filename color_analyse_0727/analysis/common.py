"""Shared constants, I/O and small helpers for the v2 analysis stages."""

from __future__ import annotations

import ast
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

import numpy as np
import pandas as pd

ROOT = Path(__file__).resolve().parents[2]
MODULE_ROOT = ROOT / "color_analyse_0727"
PROCESS_ROOT = MODULE_ROOT / "process_data"
SOURCE_METADATA_ROOT = ROOT / "processed_data"
STIMULI_ROOT = ROOT / "visual_experiment" / "stimuli_pic"

SUBJECTS: tuple[str, ...] = tuple(f"test00{i}" for i in range(1, 8))
FS = 500.0
RADIUS_MM = 20.0

ROI_TERMS = ("fusiform", "temporal_inf", "temporal_mid")

TASK1_PAIRS = (
    ("face", "face_color", "face_gray"),
    ("object", "object_color", "object_gray"),
    ("body", "body_color", "body_gray"),
    ("place", "place_color", "place_gray"),
)
GRAY_FRUITS = ("cabbage", "kiwi", "strawberry", "watermelon")
RED_FRUITS = ("strawberry", "watermelon")
GREEN_FRUITS = ("cabbage", "kiwi")

PATCH_TALAIRACH = {
    "PC": np.array([-32.0, -76.0, -7.0]),
    "CC": np.array([-25.0, -54.0, -10.0]),
    "AC": np.array([-32.0, -37.0, -8.0]),
}

# Window catalogue. The current primary window is 100-400 ms, chosen for the
# completed-stimulus/early-memory interval. Other windows remain available for
# explicitly requested sensitivity analyses.
WINDOWS: tuple[tuple[float, float], ...] = (
    (0.0, 300.0),
    (1.0, 300.0),
    (50.0, 350.0),
    (51.0, 350.0),
    (100.0, 300.0),
    (100.0, 400.0),
    (150.0, 450.0),
)
WINDOW_LABELS: tuple[str, ...] = (
    "0-300",
    "1-300",
    "50-350",
    "51-350",
    "100-300",
    "100-400",
    "150-450",
)
DEFAULT_WINDOW_LABELS: tuple[str, ...] = ("100-400",)
SIGNALS: tuple[str, ...] = ("lf30", "raw200")
SIGNAL_BANDS: dict[str, tuple[float, float] | None] = {
    "lf30": (1.0, 30.0),
    "raw200": None,  # HDF5 epochs are already 1-200 Hz filtered
}
SIGNAL_LABELS: dict[str, str] = {
    "lf30": "low-frequency 1-30 Hz",
    "raw200": "raw 1-200 Hz (HDF5 epochs)",
}
BASELINE_MS: tuple[float, float] = (-200.0, 0.0)


@dataclass(frozen=True)
class AnalysisVariant:
    window: tuple[float, float]
    signal: str

    def __post_init__(self) -> None:
        if self.window not in WINDOWS:
            raise ValueError(f"Unsupported window: {self.window}")
        if self.signal not in SIGNALS:
            raise ValueError(f"Unsupported signal: {self.signal}")

    @property
    def window_label(self) -> str:
        return f"{int(self.window[0])}-{int(self.window[1])}"

    @property
    def signal_label(self) -> str:
        return SIGNAL_LABELS[self.signal]

    @property
    def signal_band(self) -> tuple[float, float]:
        return SIGNAL_BANDS[self.signal]

    @property
    def suffix(self) -> str:
        return f"{self.window_label}_{self.signal}"


def all_variants() -> list[AnalysisVariant]:
    return [AnalysisVariant(window, signal) for window in WINDOWS for signal in SIGNALS]


def tal2mni_brett(tal: Iterable[float]) -> np.ndarray:
    """Convert Talairach to MNI with the inverse Brett piecewise affine map."""

    x, y, z = np.asarray(tuple(tal), dtype=float)
    if z < 0:
        mni_to_tal = np.array(
            [[0.9900, 0.0, 0.0], [0.0, 0.9688, 0.0420], [0.0, -0.0485, 0.8390]],
            dtype=float,
        )
    else:
        mni_to_tal = np.array(
            [[0.9900, 0.0, 0.0], [0.0, 0.9688, 0.0460], [0.0, -0.0485, 0.9189]],
            dtype=float,
        )
    return np.linalg.solve(mni_to_tal, np.array([x, y, z], dtype=float))


def patch_table() -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for patch, tal in PATCH_TALAIRACH.items():
        mni_left = tal2mni_brett(tal)
        mni_right = mni_left.copy()
        mni_right[0] *= -1.0
        rows.extend(
            [
                {
                    "patch": patch,
                    "hemisphere": "L",
                    "coordinate_space_input": "Talairach",
                    "talairach_x": tal[0],
                    "talairach_y": tal[1],
                    "talairach_z": tal[2],
                    "mni_x": mni_left[0],
                    "mni_y": mni_left[1],
                    "mni_z": mni_left[2],
                    "transform": "Brett piecewise affine inverse; z<0 branch",
                },
                {
                    "patch": patch,
                    "hemisphere": "R",
                    "coordinate_space_input": "Talairach mirrored",
                    "talairach_x": -tal[0],
                    "talairach_y": tal[1],
                    "talairach_z": tal[2],
                    "mni_x": mni_right[0],
                    "mni_y": mni_right[1],
                    "mni_z": mni_right[2],
                    "transform": "Brett piecewise affine inverse; x mirrored",
                },
            ]
        )
    return pd.DataFrame(rows)


def decode_scalar(value: Any) -> str:
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    return str(value)


def parse_coord(value: Any) -> np.ndarray | None:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return None
    try:
        parsed = ast.literal_eval(str(value))
        arr = np.asarray(parsed, dtype=float).reshape(-1)
        return arr if arr.size == 3 and np.isfinite(arr).all() else None
    except Exception:
        nums = re.findall(r"[-+]?\d*\.?\d+", str(value))
        if len(nums) == 3:
            arr = np.asarray([float(x) for x in nums], dtype=float)
            return arr if np.isfinite(arr).all() else None
    return None


def read_localization(subject: str) -> pd.DataFrame:
    from openpyxl import load_workbook

    path = SOURCE_METADATA_ROOT / subject / f"{subject}_ieegloc.xlsx"
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(v).strip() if v is not None else "" for v in next(rows)]
    records: list[dict[str, Any]] = []
    for row in rows:
        vals = list(row)
        if not vals or not vals[0]:
            continue
        rec = {header[i]: vals[i] if i < len(vals) else None for i in range(len(header))}
        channel = str(rec.get("Channel", "")).strip().upper().replace(" ", "")
        coord = parse_coord(rec.get("MNI"))
        atlas_text = "|".join(
            str(rec.get(h, ""))
            for h in header
            if any(
                token in h.lower()
                for token in (
                    "aal3",
                    "dkt",
                    "desikan",
                    "destrieux",
                    "julich",
                    "bn_atlas",
                    "brodmann",
                )
            )
        ).lower().replace(" ", "_").replace("-", "_")
        roi_hits = [term for term in ROI_TERMS if term in atlas_text]
        records.append(
            {
                "subject": subject,
                "channel": channel,
                "mni": coord,
                "mni_x": float(coord[0]) if coord is not None else np.nan,
                "mni_y": float(coord[1]) if coord is not None else np.nan,
                "mni_z": float(coord[2]) if coord is not None else np.nan,
                "roi": ";".join(roi_hits),
                "is_target_roi": bool(roi_hits),
                "color_with_sti": bool(rec.get("color_with_sti", False)),
                "stim_behavior_recorded": bool(rec.get("stim_behavior_recorded", False)),
                "atlas_text": atlas_text,
            }
        )
    wb.close()
    table = pd.DataFrame(records)
    if table.empty or not table["channel"].duplicated().any():
        return table
    merged: list[dict[str, Any]] = []
    for channel, grp in table.groupby("channel", sort=False):
        first = grp.iloc[0].to_dict()
        for field in ("mni", "mni_x", "mni_y", "mni_z", "roi", "atlas_text"):
            if field in {"roi", "atlas_text"}:
                values = [
                    str(v)
                    for v in grp[field].tolist()
                    if str(v) not in {"", "nan", "None"}
                ]
                first[field] = ";".join(dict.fromkeys(";".join(values).split(";")))
            else:
                valid = [
                    v
                    for v in grp[field].tolist()
                    if v is not None and not (isinstance(v, float) and np.isnan(v))
                ]
                first[field] = valid[0] if valid else (np.nan if field != "mni" else None)
        first["is_target_roi"] = bool(grp["is_target_roi"].any())
        first["color_with_sti"] = bool(grp["color_with_sti"].any())
        first["stim_behavior_recorded"] = bool(grp["stim_behavior_recorded"].any())
        merged.append(first)
    return pd.DataFrame(merged)


def h5_path(subject: str, task: int) -> Path:
    return PROCESS_ROOT / subject / f"task{task}_epoched_1_200Hz.h5"


def h5_labels(subject: str, task: int) -> list[str]:
    import h5py

    with h5py.File(h5_path(subject, task), "r") as h5:
        return [decode_scalar(x).strip().upper() for x in h5["labels"][()]]


def natural_key(label: str) -> tuple[str, int, str]:
    label = str(label).upper()
    match = re.fullmatch(r"([A-Z]+)(\d+)", label)
    if match:
        return match.group(1), int(match.group(2)), label
    if label.isdigit():
        return "ZZZ", int(label), label
    return label, -1, label


def common_channels(subject: str) -> list[str]:
    sets = [set(h5_labels(subject, task)) for task in (1, 2, 3)]
    return sorted(set.intersection(*sets), key=natural_key)


def load_conditions(
    subject: str,
    task: int,
    conditions: Iterable[str],
    channels: Iterable[str] | None = None,
) -> tuple[dict[str, np.ndarray], np.ndarray, list[str]]:
    import h5py

    wanted = list(conditions)
    with h5py.File(h5_path(subject, task), "r") as h5:
        labels = [decode_scalar(x).strip().upper() for x in h5["labels"][()]]
        indices = (
            list(range(len(labels)))
            if channels is None
            else [labels.index(str(c).upper()) for c in channels]
        )
        time_ms = np.asarray(h5["time_ms"][()], dtype=float)
        result = {
            name: np.asarray(h5["epochs"][name][:, indices, :], dtype=np.float32)
            for name in wanted
        }
    return result, time_ms, [labels[i] for i in indices]


def baseline_subtract(
    data: np.ndarray,
    time_ms: np.ndarray,
    start: float = BASELINE_MS[0],
    end: float = BASELINE_MS[1],
) -> np.ndarray:
    idx = (time_ms >= start) & (time_ms <= end)
    base = np.nanmean(data[..., idx], axis=-1, keepdims=True)
    return data - base


def bh_adjust(p_values: Iterable[float]) -> np.ndarray:
    p = np.asarray(list(p_values), dtype=float)
    out = np.full_like(p, np.nan)
    finite = np.isfinite(p)
    if not finite.any():
        return out
    vals = p[finite]
    order = np.argsort(vals)
    ranked = vals[order]
    adjusted = ranked * len(ranked) / np.arange(1, len(ranked) + 1)
    adjusted = np.minimum.accumulate(adjusted[::-1])[::-1]
    tmp = np.empty_like(vals)
    tmp[order] = np.minimum(adjusted, 1.0)
    out[finite] = tmp
    return out


def df_to_markdown(df: pd.DataFrame | None) -> str:
    """Render a DataFrame as a GitHub-flavored markdown table.

    Avoids the optional ``tabulate`` dependency required by pandas 3.x
    ``to_markdown``, so the pipeline runs in any environment.
    """
    if df is None or df.empty:
        return "_（空表）_"

    def fmt(value: object) -> str:
        if isinstance(value, float):
            if np.isnan(value):
                return ""
            return f"{value:.4g}"
        return str(value)

    columns = [str(c) for c in df.columns]
    lines = ["| " + " | ".join(columns) + " |", "|" + "|".join(["---"] * len(columns)) + "|"]
    for record in df.itertuples(index=False, name=None):
        lines.append("| " + " | ".join(fmt(v) for v in record) + " |")
    return "\n".join(lines)
