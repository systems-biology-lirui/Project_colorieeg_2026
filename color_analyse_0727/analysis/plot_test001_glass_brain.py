"""Plot test001 PCA time-frequency decoding on a nilearn glass brain."""

from __future__ import annotations

import argparse
import ast
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from nilearn import plotting
from openpyxl import load_workbook

from analysis.common import common_channels


ANALYSES = (
    "task3_within_pca_timefreq",
    "task2_cross_fruit_pca_timefreq",
    "task3_to_task2_pca_timefreq",
    "task2_to_task3_pca_timefreq",
)


def _localization(subject: str, root: Path, channels: list[str]) -> np.ndarray:
    path = root / "processed_data" / subject / f"{subject}_ieegloc.xlsx"
    sheet = load_workbook(path, read_only=True, data_only=True).active
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(value).strip() if value is not None else "" for value in rows[0]]
    channel_index = header.index("Channel")
    mni_index = header.index("MNI")
    values: dict[str, np.ndarray] = {}
    for row in rows[1:]:
        if channel_index >= len(row) or not row[channel_index]:
            continue
        label = str(row[channel_index]).strip().upper().replace(" ", "")
        try:
            coordinate = np.asarray(ast.literal_eval(str(row[mni_index])), dtype=float).reshape(-1)
        except (ValueError, SyntaxError, TypeError):
            continue
        if coordinate.size == 3 and np.isfinite(coordinate).all():
            values[label] = coordinate
    missing = [channel for channel in channels if channel.upper() not in values]
    if missing:
        raise RuntimeError(f"Missing MNI coordinates for {subject}: {missing}")
    return np.stack([values[channel.upper()] for channel in channels], axis=0)


def plot_glass_brain(result_dir: Path, root: Path, subject: str, output: Path) -> None:
    summary = pd.read_csv(result_dir / "electrode_summary.csv")
    summary = summary[summary["subject"].astype(str) == subject]
    channels = common_channels(subject)
    if summary["channel"].nunique() != len(channels):
        raise RuntimeError(
            f"Expected {len(channels)} channels in summary, found {summary['channel'].nunique()}"
        )
    coordinates = _localization(subject, root, channels)
    figure = plt.figure(figsize=(16, 10), facecolor="white")
    peak_values_all = summary["peak_accuracy"].to_numpy(dtype=float)
    size_min = float(np.nanmin(peak_values_all))
    size_max = float(np.nanmax(peak_values_all))

    def peak_sizes(values: np.ndarray) -> np.ndarray:
        if size_max <= size_min:
            return np.full(values.shape, 70.0, dtype=float)
        scaled = (values - size_min) / (size_max - size_min)
        return 28.0 + 125.0 * np.clip(scaled, 0.0, 1.0)

    for panel_index, analysis in enumerate(ANALYSES):
        panel = summary[summary["analysis"] == analysis].set_index("channel").loc[channels]
        significant = panel["n_clusters_p_le_0.05"].to_numpy(dtype=float) > 0
        if not significant.any():
            raise RuntimeError(f"No significant-cluster electrodes for {subject} {analysis}")
        panel = panel.iloc[np.flatnonzero(significant)]
        panel_coordinates = coordinates[significant]
        peak_times = panel["peak_time_ms"].to_numpy(dtype=float)
        peak_values = panel["peak_accuracy"].to_numpy(dtype=float)
        axis = figure.add_subplot(2, 2, panel_index + 1)
        plotting.plot_markers(
            peak_times,
            panel_coordinates,
            node_size=peak_sizes(peak_values),
            node_cmap="RdBu",
            node_vmin=0.0,
            node_vmax=800.0,
            display_mode="ortho",
            figure=figure,
            axes=axis,
            title=analysis.replace("_", " "),
            annotate=True,
            colorbar=(panel_index == 3),
            node_kwargs={"edgecolor": "white", "linewidth": 0.35},
        )
    transform_label = "no PCA" if "no_pca" in result_dir.name.lower() else "PCA(10)"
    figure.suptitle(
        f"{subject} all-electrode {transform_label} time-frequency decoding | nilearn glass brain\n"
        "only electrodes with p≤0.05 time clusters; color = peak latency (early red → late blue); marker size = peak accuracy",
        fontsize=14,
    )
    figure.text(0.5, 0.02, "MNI coordinates; orthogonal sagittal/coronal/axial views; point-size range is scaled across all four branches", ha="center", fontsize=10)
    figure.tight_layout(rect=(0, 0.04, 1, 0.94))
    output.parent.mkdir(parents=True, exist_ok=True)
    figure.savefig(output, dpi=220, bbox_inches="tight")
    plt.close(figure)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--result-dir", type=Path, required=True)
    parser.add_argument("--subject", default="test001")
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()
    root = Path(__file__).resolve().parents[2]
    plot_glass_brain(args.result_dir.resolve(), root, args.subject, args.output.resolve())
    print(args.output.resolve())


if __name__ == "__main__":
    main()
